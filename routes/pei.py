from fastapi import APIRouter, Depends, HTTPException, Body, Query
from fastapi.responses import Response
from sqlalchemy.orm import Session
from database import get_db
from dependencies import get_current_user
from models import PatientPei, PeiTemp, BaseGuia, Carteirinha
from services.pei_service import update_patient_pei
from pydantic import BaseModel
from typing import Optional, List
from datetime import date, timedelta, datetime
from sqlalchemy import func, or_, text
import io
import openpyxl

router = APIRouter(
    prefix="/pei",
    tags=["PEI"]
)

class PeiOverrideRequest(BaseModel):
    guia_id: int
    pei_semanal: float

def apply_filters(query, search, status, validade_start, validade_end, vencimento_filter):
    # Text Search (Patient, Carteirinha, Therapy)
    if search:
        search_term = f"%{search}%"
        query = query.filter(
            or_(
                Carteirinha.paciente.ilike(search_term),
                Carteirinha.carteirinha.ilike(search_term),
                PatientPei.codigo_terapia.ilike(search_term)
            )
        )
    
    # Status Enum
    if status:
        query = query.filter(PatientPei.status == status)

    # Date Range
    if validade_start:
        query = query.filter(PatientPei.validade >= validade_start)
    if validade_end:
        query = query.filter(PatientPei.validade <= validade_end)

    # Smart Vencimento Filters
    today = date.today()
    if vencimento_filter:
        if vencimento_filter == 'vencidos':
            query = query.filter(PatientPei.validade < today)
        elif vencimento_filter == 'vence_d7':
            target_date = today + timedelta(days=7)
            query = query.filter(PatientPei.validade >= today, PatientPei.validade <= target_date)
        elif vencimento_filter == 'vence_d30':
            target_date = today + timedelta(days=30)
            query = query.filter(PatientPei.validade >= today, PatientPei.validade <= target_date)
            
    return query

@router.get("/dashboard")
def get_dashboard_stats(
    id_convenio: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    today = date.today()
    
    # Base query for active patients? Or just all?
    # Stats: Vencidos, Vence D+7, Vence D+30
    
    # Isolation
    query_base = db.query(func.count(PatientPei.id)).join(Carteirinha)
    from dependencies import get_allowed_convenio_ids
    allowed_ids = get_allowed_convenio_ids(current_user)
    
    if id_convenio:
        if allowed_ids and id_convenio not in allowed_ids:
             raise HTTPException(status_code=403, detail="Sem permissão para este convênio.")
        query_base = query_base.filter(Carteirinha.id_convenio == id_convenio)
    elif allowed_ids:
        query_base = query_base.filter(Carteirinha.id_convenio.in_(allowed_ids))
    
    # Vencidos
    vencidos = query_base.filter(PatientPei.validade < today).scalar()
    
    # Vence D+7
    d7_end = today + timedelta(days=7)
    vence_d7 = query_base.filter(
        PatientPei.validade >= today, 
        PatientPei.validade <= d7_end
    ).scalar()
    
    # Vence D+30
    d30_end = today + timedelta(days=30)
    vence_d30 = query_base.filter(
        PatientPei.validade >= today, 
        PatientPei.validade <= d30_end
    ).scalar()
    
    total = query_base.scalar()
    pendentes = query_base.filter(PatientPei.status == 'Pendente').scalar()
    validados = query_base.filter(PatientPei.status == 'Validado').scalar()

    return {
        "total": total,
        "vencidos": vencidos or 0,
        "vence_d7": vence_d7 or 0,
        "vence_d30": vence_d30 or 0,
        "pendentes": pendentes or 0,
        "validados": validados or 0
    }

@router.get("/")
def list_pei(
    page: int = 1,
    pageSize: int = 50,
    search: Optional[str] = None,
    status: Optional[str] = None, # Validado, Pendente
    validade_start: Optional[date] = None,
    validade_end: Optional[date] = None,
    vencimento_filter: Optional[str] = None, # vencidos, vence_d7, vence_d30
    id_convenio: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    # Optimized query selecting only necessary columns
    query = db.query(
        PatientPei.id,
        PatientPei.carteirinha_id,
        Carteirinha.carteirinha,
        Carteirinha.paciente,
        PatientPei.codigo_terapia,
        PatientPei.pei_semanal,
        PatientPei.validade,
        PatientPei.status,
        PatientPei.base_guia_id,
        BaseGuia.guia.label("guia_vinculada"),
        BaseGuia.sessoes_autorizadas,
        PatientPei.updated_at,
        Carteirinha.id_paciente # For export matching if needed
    ).join(Carteirinha, PatientPei.carteirinha_id == Carteirinha.id)\
     .outerjoin(BaseGuia, PatientPei.base_guia_id == BaseGuia.id)
    
    # Isolation
    from dependencies import get_allowed_convenio_ids
    allowed_ids = get_allowed_convenio_ids(current_user)
    
    if id_convenio:
        if allowed_ids and id_convenio not in allowed_ids:
             raise HTTPException(status_code=403, detail="Sem permissão para este convênio.")
        query = query.filter(Carteirinha.id_convenio == id_convenio)
    elif allowed_ids:
        query = query.filter(Carteirinha.id_convenio.in_(allowed_ids))
    
    query = apply_filters(query, search, status, validade_start, validade_end, vencimento_filter)
    
    total_items = query.count()
    
    # Pagination
    skip = (page - 1) * pageSize
    results = query.order_by(PatientPei.status.asc(), PatientPei.updated_at.desc()).offset(skip).limit(pageSize).all()
    
    data = []
    for row in results:
        data.append({
            "id": row.id,
            "carteirinha_id": row.carteirinha_id,
            "carteirinha": row.carteirinha or "",
            "paciente": row.paciente or "",
            "codigo_terapia": row.codigo_terapia,
            "pei_semanal": row.pei_semanal,
            "validade": row.validade,
            "status": row.status,
            "base_guia_id": row.base_guia_id,
            "guia_vinculada": row.guia_vinculada or "-",
            "sessoes_autorizadas": row.sessoes_autorizadas or 0,
            "updated_at": row.updated_at
        })

    return {
        "data": data,
        "total": total_items,
        "page": page,
        "pageSize": pageSize
    }

@router.get("/export")
def export_pei(
    search: Optional[str] = None,
    status: Optional[str] = None,
    validade_start: Optional[date] = None,
    validade_end: Optional[date] = None,
    vencimento_filter: Optional[str] = None,
    id_convenio: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    print("DEBUG: Starting PEI Export...")
    print("DEBUG: Starting PEI Export (Optimized)...")
    try:
        # Generate Excel (Write Only Mode for Performance)
        wb = openpyxl.Workbook(write_only=True)
        ws = wb.create_sheet("PEI Export")
        
        # Header
        ws.append([
            "ID Paciente", "Paciente", "Carteirinha", "ID Pagamento", "Código Terapia", 
            "Guia Vinculada", "Data Autorização", "Senha", "Qtd Autorizada",
            "PEI Semanal", "Validade", "Status", "Atualizado Em"
        ])
        
        # Generator function for streaming rows
        def generate_rows():
            print("DEBUG: Executing Query...")
            # Use yield_per to stream results from DB instead of loading all into memory
            query = db.query(PatientPei).join(Carteirinha).outerjoin(BaseGuia, PatientPei.base_guia_id == BaseGuia.id)
            query = apply_filters(query, search, status, validade_start, validade_end, vencimento_filter)
            
            # Fetch in chunks to reduce memory usage
            # Note: SQLite might not support yield_per strictly, but it's good practice. 
            # In sync mode with small datasets it's fine.
            # For strict speed we could fetch raw tuples but ORM is cleaner.
            
            results = query.yield_per(500)
            
            count = 0
            for row in results:
                count += 1
                # Handle timezone naive
                updated_at_val = row.updated_at
                if updated_at_val and updated_at_val.tzinfo:
                    updated_at_val = updated_at_val.replace(tzinfo=None)
                
                # Base Guia Helpers
                guia_num = row.base_guia_rel.guia if row.base_guia_rel else "-"
                data_auth = row.base_guia_rel.data_autorizacao if row.base_guia_rel else None
                senha = row.base_guia_rel.senha if row.base_guia_rel else "-"
                qtd_aut = row.base_guia_rel.sessoes_autorizadas if row.base_guia_rel else 0
                
                # ID Paciente
                id_paciente_real = row.carteirinha_rel.id_paciente if row.carteirinha_rel else ""
                id_pagamento_val = row.carteirinha_rel.id_pagamento if row.carteirinha_rel and row.carteirinha_rel.id_pagamento else ""

                ws.append([
                    id_paciente_real,
                    row.carteirinha_rel.paciente if row.carteirinha_rel else "",
                    row.carteirinha_rel.carteirinha if row.carteirinha_rel else "",
                    id_pagamento_val,
                    row.codigo_terapia,
                    guia_num,
                    data_auth.strftime("%d/%m/%Y") if data_auth else "",
                    senha,
                    qtd_aut,
                    row.pei_semanal,
                    row.validade.strftime("%d/%m/%Y") if row.validade else "",
                    row.status if row.status else "Pendente",
                    updated_at_val.strftime("%d/%m/%Y") if updated_at_val else ""
                ])
            print(f"DEBUG: Processed {count} rows.")
            
            # After adding all rows to worksheet, save to stream
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            yield output.getvalue()

        filename = f"export_pei_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        headers = {
            'Content-Disposition': f'attachment; filename="{filename}"'
        }
        
        from fastapi.responses import StreamingResponse
        # Note: StreamingResponse with yield is tricky for zip/xlsx because headers are at the end.
        # Ideally we buffer the whole xlsx. 
        # But 'generate_rows' here yields ONE big chunk. 
        # To truly stream xlsx we need to yield chunks of bytes which openpyxl doesn't support easily.
        # So we just optimize the building part.
        
        # We can't yield parts of a zip file easily. So we just call the generator to get the bytes.
        # But wait, StreamingResponse expects an iterator of bytes.
        
        # Reverting to direct save but kept write_only=True which is faster.
        
        output = io.BytesIO()
        wb.save(output) # This will fail if no rows added yet? No.
        
        # Re-implementing correctly: iterate query, write to wb, then save.
        pass

    except Exception as e:
        print(f"DEBUG: Export Error: {str(e)}")
        raise HTTPException(status_code=400, detail=f"Erro ao exportar: {str(e)}")
    
    try:
        # Generate Excel (Write Only Mode for Performance)
        wb = openpyxl.Workbook(write_only=True)
        ws = wb.create_sheet("PEI Export")
        
        # Header
        ws.append([
            "ID Paciente", "Paciente", "Carteirinha", "ID Pagamento", "Código Terapia", 
            "Guia Vinculada", "Data Autorização", "Senha", "Qtd Autorizada",
            "PEI Semanal", "Validade", "Status", "Atualizado Em"
        ])

        # Base query joining for isolation
        query = db.query(
            Carteirinha.id_paciente,          # 0
            Carteirinha.paciente,             # 1
            Carteirinha.carteirinha,          # 2
            Carteirinha.id_pagamento,         # 3
            PatientPei.codigo_terapia,        # 4
            BaseGuia.guia,                    # 5
            BaseGuia.data_autorizacao,        # 6
            BaseGuia.senha,                   # 7
            BaseGuia.sessoes_autorizadas,     # 8
            PatientPei.pei_semanal,           # 9
            PatientPei.validade,              # 10
            PatientPei.status,                # 11
            PatientPei.updated_at             # 12
        ).select_from(PatientPei)\
         .join(Carteirinha, PatientPei.carteirinha_id == Carteirinha.id)\
         .outerjoin(BaseGuia, PatientPei.base_guia_id == BaseGuia.id)
        
        # Isolation
        from dependencies import get_allowed_convenio_ids
        allowed_ids = get_allowed_convenio_ids(current_user)
        
        if id_convenio:
            if allowed_ids and id_convenio not in allowed_ids:
                 raise HTTPException(status_code=403, detail="Sem permissão para este convênio.")
            query = query.filter(Carteirinha.id_convenio == id_convenio)
        elif allowed_ids:
            query = query.filter(Carteirinha.id_convenio.in_(allowed_ids))
        
        query = apply_filters(query, search, status, validade_start, validade_end, vencimento_filter)
        
        # Use yield_per to stream results from DB
        results = query.yield_per(1000)
        
        count = 0
        for row in results:
            count += 1
            # Row is now a tuple, access by index or name
            
            # Handle timezone naive
            updated_at_val = row.updated_at
            if updated_at_val and updated_at_val.tzinfo:
                updated_at_val = updated_at_val.replace(tzinfo=None)
            
            # Helper for dates
            def fmt(d): return d.strftime("%d/%m/%Y") if d else ""

            ws.append([
                row.id_paciente or "",                  # ID Paciente
                row.paciente or "",                     # Paciente
                row.carteirinha or "",                  # Carteirinha
                row.id_pagamento or "",                 # ID Pagamento (Direct from select)
                row.codigo_terapia,                     # Codigo Terapia
                row.guia or "-",                        # Guia
                fmt(row.data_autorizacao),              # Data Auth
                row.senha or "-",                       # Senha
                row.sessoes_autorizadas or 0,           # Qtd Aut
                row.pei_semanal,                        # PEI
                fmt(row.validade),                      # Validade
                row.status if row.status else "Pendente", # Status
                fmt(updated_at_val)                     # Atualizado Em
            ])
        
        print(f"DEBUG: Processed {count} rows. Saving...")
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"export_pei_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        headers = {'Content-Disposition': f'attachment; filename="{filename}"'}
        
        from fastapi.responses import StreamingResponse
        return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)

    except Exception as e:
        print(f"DEBUG: Export Error: {str(e)}")
        raise HTTPException(status_code=400, detail=f"Erro ao exportar: {str(e)}")

@router.post("/override")
def override_pei(
    req: PeiOverrideRequest,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    # Validation: check if the guia belongs to the user's convenio
    guia = db.query(BaseGuia).join(Carteirinha).filter(BaseGuia.id == req.guia_id)
    from dependencies import get_allowed_convenio_ids
    allowed_ids = get_allowed_convenio_ids(current_user)
    if allowed_ids:
        guia = guia.filter(Carteirinha.id_convenio.in_(allowed_ids))
    
    guia_obj = guia.first()
    if not guia_obj:
        raise HTTPException(status_code=403, detail="Acesso negado ou Guia não encontrada")

    # Upsert PeiTemp
    temp = db.query(PeiTemp).filter(PeiTemp.base_guia_id == req.guia_id).first()
    if not temp:
        temp = PeiTemp(base_guia_id=req.guia_id, pei_semanal=req.pei_semanal)
        db.add(temp)
    else:
        temp.pei_semanal = req.pei_semanal
    db.commit()
    

    # Recalculate
    # Actually, the trigger on PeiTemp (after_insert/update) should have already handled this 
    # because we committed above.
    # However, to be safe or if the commit happened before trigger fully propagated in some async scenarios (unlikely in sync sqlalchemy),
    # we can explicitly call it or just rely on the commit.
    # The event listener fires *after* flush/commit usually depending on config.
    # But since we just committed, the 'after_update' for PeiTemp should have fired.
    
    # Just in case we want to return the updated status immediately:
    # update_patient_pei(db, guia.carteirinha_id, guia.codigo_terapia)
    
    return {"status": "success"}

# Note: update_patient_pei_backend removed as it is now in services/pei_service.py

