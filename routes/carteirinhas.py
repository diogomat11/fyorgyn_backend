from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Body
from sqlalchemy.orm import Session
from database import get_db
from models import Carteirinha, Job, BaseGuia, User
from typing import List, Optional
import io
import csv
from openpyxl import load_workbook
from sqlalchemy import or_, String, cast
from dependencies import get_current_user

router = APIRouter(
    prefix="/carteirinhas",
    tags=["Carteirinhas"]
)

def validate_carteirinha_format(code: str):
    # Removida validação rígida de 21 caracteres para suportar outros convênios (ex: IPASGO, Amil, etc).
    # Apenas passa, confiando na importação/input do usuário.
    pass

def normalize_header(header):
    header = str(header).strip()
    mapping = {
        'carteiras': 'Carteirinha',
        'Carteiras': 'Carteirinha',
        'carteirinha': 'Carteirinha',
        'Carteirinha': 'Carteirinha',
        'PACIENTE': 'Paciente',
        'paciente': 'Paciente',
        'Paciente': 'Paciente',
        'ID': 'IdPaciente',
        'id': 'IdPaciente',
        'IdPaciente': 'IdPaciente',
        'id_paciente': 'IdPaciente',
        'Codigo_beneficiario': 'Codigo_beneficiario',
        'codigo_beneficiario': 'Codigo_beneficiario',
        'id_convenio': 'id_convenio',
        'IdConvenio': 'id_convenio',
        'status': 'status',
        'Status': 'status',
        'STATUS': 'status'
    }
    return mapping.get(header, header)

@router.post("/upload")
async def upload_carteirinhas(
    file: UploadFile = File(...),
    overwrite: bool = Form(False),
    id_convenio: Optional[int] = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    from dependencies import get_allowed_convenio_ids
    allowed_ids = get_allowed_convenio_ids(user)
    target_convenio = id_convenio if id_convenio else (allowed_ids[0] if allowed_ids else None)
    
    try:
        contents = await file.read()
        rows = []
        
        # 1. Parse File into List of Dicts
        if file.filename.endswith('.csv'):
            # Try decoding with utf-8-sig (handles BOM) then latin1
            text_content = ""
            try:
                text_content = contents.decode('utf-8-sig')
            except UnicodeDecodeError:
                text_content = contents.decode('latin1')
            
            # Detect separator
            dialect = 'excel' # default comma
            if ';' in text_content and text_content.count(';') > text_content.count(','):
                class SemiColonDialect(csv.Dialect):
                    delimiter = ';'
                    quotechar = '"'
                    doublequote = True
                    skipinitialspace = False
                    lineterminator = '\r\n'
                    quoting = csv.QUOTE_MINIMAL
                dialect = SemiColonDialect
            
            f = io.StringIO(text_content)
            reader = csv.DictReader(f, dialect=dialect)
            
            # Normalize Headers
            # DictReader uses fieldnames from first row. We need to remap them.
            if reader.fieldnames:
                original_headers = reader.fieldnames
                normalized_fieldnames = [normalize_header(h) for h in original_headers]
                reader.fieldnames = normalized_fieldnames # Override fieldnames
            
            for row in reader:
                rows.append(row)
                
        else:
            # Excel (.xlsx)
            wb = load_workbook(filename=io.BytesIO(contents), read_only=True, data_only=True)
            ws = wb.active
            
            header_map = {} # col_index -> normalized_name
            is_header = True
            
            for row in ws.iter_rows(values_only=True):
                if not row: continue
                
                # Check for empty row
                if all(cell is None for cell in row): continue

                if is_header:
                    for i, cell_value in enumerate(row):
                        if cell_value:
                            header_map[i] = normalize_header(cell_value)
                    is_header = False
                    continue
                
                # Data Row
                row_data = {}
                for i, cell_value in enumerate(row):
                    if i in header_map:
                        row_data[header_map[i]] = cell_value
                
                rows.append(row_data)
            
            wb.close()
        
        # 2. Process Rows
        carteirinhas_data = []
        errors = []
        
        # Check required columns
        if rows:
             first_row_keys = rows[0].keys()
             if 'Carteirinha' not in first_row_keys:
                  raise HTTPException(status_code=400, detail=f"Arquivo inválido. Coluna 'Carteirinha' não encontrada. Colunas encontradas: {list(first_row_keys)}")
        else:
             pass 

        from models import Convenio
        convenios_db = db.query(Convenio).all()
        valid_convenios = {c.id_convenio for c in convenios_db}
        convenio_digits_map = {c.id_convenio: c.digitos_carteirinha for c in convenios_db if c.digitos_carteirinha}

        for index, row in enumerate(rows):
            # Set dynamic fallback if available in column, else use default Form id_convenio
            target_convenio = id_convenio
            cart_raw = row.get('Carteirinha')
            cart = str(cart_raw).strip() if cart_raw is not None else ""
            
            paciente_raw = row.get('Paciente')
            paciente = str(paciente_raw).strip() if paciente_raw is not None else ""
            
            # Convert IDs to integers
            id_paciente = None
            id_pagamento = None
            
            if 'IdPaciente' in row and row['IdPaciente']:
                try:
                    val = str(row['IdPaciente']).strip()
                    if val and val.lower() != 'nan' and val.lower() != 'none':
                        id_paciente = int(float(val))
                except (ValueError, TypeError):
                    pass
            
            if 'id_convenio' in row and row['id_convenio']:
                try:
                    val = str(row['id_convenio']).strip()
                    if val and val.lower() != 'nan' and val.lower() != 'none':
                        target_convenio = int(float(val))
                except (ValueError, TypeError):
                    pass
            
            codigo_beneficiario = str(row.get('Codigo_beneficiario', '')).strip()
            if codigo_beneficiario.lower() == 'nan' or codigo_beneficiario.lower() == 'none' or not codigo_beneficiario:
                codigo_beneficiario = None
            
            
            status_val = row.get('status', 'ativo')
            if not status_val or str(status_val).lower() == 'nan':
                status_val = 'ativo'

            if cart and cart.lower() != 'nan' and cart.lower() != 'none':
                # Dynamic format validation
                req_len = convenio_digits_map.get(target_convenio)
                if req_len:
                    digits_only = ''.join(filter(str.isdigit, cart))
                    if len(digits_only) != req_len:
                         errors.append(f"Linha {index+2}: Carteirinha {cart} inválida para este convênio. Esperado {req_len} dígitos, recebeu {len(digits_only)}.")
                         continue
                
                carteirinhas_data.append({
                    "carteirinha": cart,
                    "paciente": paciente,
                    "id_paciente": id_paciente,
                    "codigo_beneficiario": codigo_beneficiario,
                    "id_convenio": target_convenio,
                    "status": status_val
                })

        warnings = errors.copy()
        
        # Dedup items in the CSV itself to avoid UniqueViolation on db.add()
        # Report duplicates with different patients
        unique_carteirinhas = {}
        conflict_carteirinhas = set()
        
        for item in carteirinhas_data:
            cart = item["carteirinha"]
            
            if cart in conflict_carteirinhas:
                continue
                
            if cart in unique_carteirinhas:
                existing_item = unique_carteirinhas[cart]
                if str(existing_item.get("id_paciente")) != str(item.get("id_paciente")):
                    warnings.append(f"Carteirinha duplicada para pacientes distintos: {cart} (Pacientes: {existing_item['paciente']} e {item['paciente']})")
                    conflict_carteirinhas.add(cart)
                    del unique_carteirinhas[cart]
                # else: same patient, silently dedup
            else:
                unique_carteirinhas[cart] = item
                
        carteirinhas_data = list(unique_carteirinhas.values())

        if not carteirinhas_data and warnings:
             raise HTTPException(status_code=400, detail="Nenhuma carteirinha válida na importação.\n" + "\n".join(warnings[:15]))
             
        count_added = 0
        count_updated = 0
        
        for item in carteirinhas_data:
            derived_convenio = item.get('id_convenio') or target_convenio
            if derived_convenio not in valid_convenios:
                derived_convenio = target_convenio # Fallback to form/default

            existing = db.query(Carteirinha).filter(
                Carteirinha.carteirinha == item['carteirinha']
            ).first()
            
            if existing:
                # Update existing record
                changed = False
                if existing.paciente != item['paciente']:
                    existing.paciente = item['paciente']
                    changed = True
                if existing.id_paciente != item['id_paciente']:
                    existing.id_paciente = item['id_paciente']
                    changed = True
                if existing.codigo_beneficiario != item['codigo_beneficiario']:
                   existing.codigo_beneficiario = item['codigo_beneficiario']
                   changed = True
                if existing.status != item['status']:
                    existing.status = item['status']
                    changed = True
                
                # Isolation update: if it didn't have a convenio or was different, set it
                if existing.id_convenio != derived_convenio and derived_convenio:
                    existing.id_convenio = derived_convenio
                    changed = True
                
                if changed:
                    count_updated += 1
            else:
                 new_cart = Carteirinha(
                     carteirinha=item['carteirinha'],
                     paciente=item['paciente'],
                     id_paciente=item.get('id_paciente'),
                     codigo_beneficiario=item.get('codigo_beneficiario'),
                     status=item.get('status', 'ativo'),
                     id_convenio=derived_convenio
                 )
                 db.add(new_cart)
                 count_added += 1
        
        db.commit()
        
        return {
            "message": "Upload processed successfully",
            "added": count_added,
            "updated": count_updated,
            "total_processed": len(carteirinhas_data),
            "warnings": warnings
        }

    except HTTPException:
        raise
    except Exception as e:
        import traceback
        import os
        tb_str = traceback.format_exc()
        trace_path = os.path.join(os.path.dirname(__file__), '..', 'upload_trace.log')
        with open(trace_path, 'a', encoding='utf-8') as f:
            f.write(f"\n--- NEW UPLOAD ERROR ---\n{tb_str}\n")
            
        print("Upload Error traced to:", trace_path)
        raise HTTPException(status_code=500, detail=f"Internal Server Error: {str(e)}")

@router.get("", response_model=None)
@router.get("/", include_in_schema=False)
def list_carteirinhas(
    skip: int = 0, 
    limit: int = 100, 
    search: Optional[str] = None, 
    status: Optional[str] = None,
    codigo_beneficiario: Optional[str] = None,
    id_convenio: Optional[int] = None,
    paciente: Optional[str] = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    query = db.query(Carteirinha)
    
    from dependencies import get_allowed_convenio_ids
    allowed_ids = get_allowed_convenio_ids(user)
    
    if id_convenio:
        if allowed_ids and id_convenio not in allowed_ids:
             raise HTTPException(status_code=403, detail="Sem permissão para este convênio.")
        query = query.filter(Carteirinha.id_convenio == id_convenio)
    elif allowed_ids:
        query = query.filter(Carteirinha.id_convenio.in_(allowed_ids))
    
    # Text Search (General)
    if search:
        search_filter = f"%{search}%"
        # Cast integer columns to string for LIKE search
        query = query.filter(
            or_(
                Carteirinha.paciente.ilike(search_filter), 
                Carteirinha.carteirinha.ilike(search_filter),
                Carteirinha.id_paciente.cast(String).ilike(search_filter),
                Carteirinha.codigo_beneficiario.ilike(search_filter)
            )
        )
        
    # Specific Filters
    if status:
        query = query.filter(Carteirinha.status == status)
        
    if codigo_beneficiario:
        query = query.filter(Carteirinha.codigo_beneficiario.ilike(f"%{codigo_beneficiario}%"))
        
    if paciente:
        query = query.filter(Carteirinha.paciente.ilike(f"%{paciente}%"))
    
    # Sort alphabetically by patient name
    query = query.order_by(Carteirinha.paciente.asc())
    
    total = query.count()
    carteirinhas = query.offset(skip).limit(limit).all()
    
    return {
        "data": carteirinhas,
        "total": total,
        "skip": skip,
        "limit": limit
    }

@router.post("", response_model=None)
@router.post("/", include_in_schema=False)
def create_carteirinha(item: dict = Body(...), db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """Create a new carteirinha"""
    if 'carteirinha' not in item:
        raise HTTPException(status_code=400, detail="Field 'carteirinha' is required")
    
    # Validate format
    validate_carteirinha_format(item['carteirinha'])
    
    # Check if already exists
    existing = db.query(Carteirinha).filter(Carteirinha.carteirinha == item['carteirinha']).first()
    if existing:
        raise HTTPException(status_code=400, detail=f"Carteirinha {item['carteirinha']} already exists")
    
    from dependencies import get_allowed_convenio_ids
    allowed_ids = get_allowed_convenio_ids(user)
    target_convenio = item.get('id_convenio') if item.get('id_convenio') else (allowed_ids[0] if allowed_ids else None)

    new_cart = Carteirinha(
        carteirinha=item['carteirinha'],
        paciente=item.get('paciente', ''),
        id_paciente=item.get('id_paciente'),
        codigo_beneficiario=item.get('codigo_beneficiario'),
        status=item.get('status', 'ativo'),
        id_convenio=target_convenio
    )
    
    db.add(new_cart)
    db.commit()
    db.refresh(new_cart)
    
    return new_cart

@router.put("/{carteirinha_id}")
def update_carteirinha(carteirinha_id: int, item: dict = Body(...), db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    cart = db.query(Carteirinha).filter(Carteirinha.id == carteirinha_id).first()
    if not cart:
        raise HTTPException(status_code=404, detail="Carteirinha not found")
    
    if 'carteirinha' in item:
        validate_carteirinha_format(item['carteirinha'])
        cart.carteirinha = item['carteirinha']
    if 'paciente' in item:
        cart.paciente = item['paciente']
    if 'id_paciente' in item:
        cart.id_paciente = item['id_paciente']
    if 'codigo_beneficiario' in item:
        cart.codigo_beneficiario = item['codigo_beneficiario']
    if 'id_convenio' in item and item['id_convenio']:
        cart.id_convenio = item['id_convenio']
    if 'status' in item:
        cart.status = item['status']
        
    db.commit()
    db.refresh(cart)
    return cart

@router.delete("/{carteirinha_id}")
def delete_carteirinha(carteirinha_id: int, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    cart = db.query(Carteirinha).filter(Carteirinha.id == carteirinha_id).first()
    if not cart:
        raise HTTPException(status_code=404, detail="Carteirinha not found")
    
    db.delete(cart)
    db.commit()
    return {"message": "Deleted successfully"}
